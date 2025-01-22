import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from utils import match_wine_names

def engine(df_off, df_on, df_remain):
    # 매칭 및 데이터 처리
    off_array = list(df_off['상 품 명'].values[~pd.isnull(df_off['상 품 명'].values)])
    on_array = list(df_on['품목명'].values[~pd.isnull(df_on['품목명'].values)])
    remain_array = list(df_remain[df_remain.columns[1]].values[~pd.isnull(df_remain[df_remain.columns[1]].values)])[1:]

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 결과 DataFrame 생성
    columns = [
        'No.', '상품소분류', '상 품 명', 'OFF 수량', 'ON 수량', 'TOTAL 수량', '기간 개월수', '월판매량 수량',
        '현재고 수량', '발주분 수량', '총재고 수량', '재고월수', '3개월후 재고', '발주희망', '발주수량 병',
        '발주수량 박스단위', '예상발주 박스', '발주확정 박스', '발주확정 병'
    ]
    df = pd.DataFrame(columns=columns)

    df['상 품 명'] = off_array
    df['OFF 수량'] = df_off['합 계'].values[1:-2]

    # ON 수량 dict
    on_dict = {
        df_on['품목명'][i]: df_on['EA'][i] 
        for i in range(len(df_on)) 
        if df_on['품목명'][i] in on_array
    }
    
    # 매칭
    result_on_off = match_wine_names(off_array, on_array)
    result_off_remain = match_wine_names(off_array, remain_array)
    
    # 재고 수량 dict
    remain_dict = {
        df_remain[df_remain.columns[1]][i]: df_remain[df_remain.columns[9]][i] 
        for i in range(len(df_remain)) 
        if df_remain[df_remain.columns[1]][i] in remain_array
    }
    
    df['ON 수량'] = [
        on_dict[result_on_off[name]] if name in result_on_off else 0
        for name in off_array
    ]
    df['현재고 수량'] = [
        remain_dict[result_off_remain[name]] if name in result_off_remain else 0
        for name in off_array
    ]

    # 1) 엑셀 파일에 수식 입력 및 저장
    output_file = 'result_with_formulas.xlsx'
    df.to_excel(output_file, index=False, engine='openpyxl')

    wb = load_workbook(output_file)
    ws = wb.active

    # 공식 입력
    for row in range(2, ws.max_row + 1):
        ws[f'F{row}'] = f"=D{row}+E{row}"   # TOTAL
        ws[f'G{row}'] = 12                 # 기간 개월수(단순 상수)
        ws[f'H{row}'] = f"=TRUNC(F{row}/G{row})"  # 월판매량(정수 처리)
        ws[f'I{row}'] = df['현재고 수량'].iloc[row-2]
        ws[f'K{row}'] = f"=I{row}"
        # L열(재고월수) = ROUND(I / (월판매량), 2)
        ws[f'L{row}'] = f"=ROUND(I{row}/H{row},2)"
        ws[f'M{row}'] = f"=MAX(K{row}-3*H{row}, 0)"
        ws[f'N{row}'] = f"=H{row}*12 - M{row}"
        ws[f'O{row}'] = f"=N{row}"
        ws[f'P{row}'] = 12
        ws[f'Q{row}'] = f"=ROUND(O{row}/P{row}, 0)"
        ws[f'R{row}'] = f"=Q{row}"
        ws[f'S{row}'] = f"=R{row}*12"

    # L열 빨간색 처리 (엑셀의 실제 수식 결과는 없으므로, 파이썬에서 직접 계산)
    for row in range(2, ws.max_row + 1):
        i_val = ws[f'I{row}'].value  # 현재고
        # H열에는 수식이 들어있지만, 여기서는 df['OFF 수량'], df['ON 수량'] 등을 통해 직접 계산하든,
        # 혹은 ws[f'H{row}'] 값(문자열)로 파싱해서 계산하든 임의로 가능.
        #
        # 여기서는 "H=TRUNC((D+E)/12)" 의 로직을 그대로 파이썬에서 구현:
        d_val = ws[f'D{row}'].value or 0  # OFF 수량
        e_val = ws[f'E{row}'].value or 0  # ON 수량
        total_monthly_sales = int((d_val + e_val)/12)  # TRUNC() 와 동일

        if total_monthly_sales != 0:
            l_val = round(i_val / total_monthly_sales, 2)
        else:
            l_val = 9999  # 0으로 나누면 매우 큰 값 또는 별도 처리

        if l_val < 4:
            ws[f'L{row}'].fill = red_fill  # 빨간색 처리

    # 합계 행 추가
    sum_row = ws.max_row + 1
    for col in range(4, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}{sum_row}'] = f"=SUM({col_letter}2:{col_letter}{sum_row-1})"

    # 컬럼 너비 자동 조정
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_file)

    return output_file


if __name__ == '__main__':
    off_file = 'off.xlsx'
    on_file = 'on.xlsx'
    remain_file = 'remain.xlsx'
    df_off = pd.read_excel(off_file)
    df_on = pd.read_excel(on_file)
    df_remain = pd.read_excel(remain_file)
    engine(df_off, df_on, df_remain)
